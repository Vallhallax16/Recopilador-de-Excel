import openpyxl

ruta_base = "D:\\Victoria"

carpetas = ["231113","231114","231115","231116"]
subcarpeta = " Rúbricas"
nexo = "Rúbrica para evaluación "
tipos_archivo = ["científica","impacto social","escalabilidad"]
extension = ".xlsx"
archivo_objetivo = "231128 Acumulado.xlsx"

celda_promedio = "D12"
celdas_objetivo = ["Q",'R',"S"]
#celdas_objetivo = [16,17,18]

index_hojas_objetivo = 1

for n_carpeta in carpetas:
#
    index_celda_obj = 0

    for tipo in tipos_archivo:
    #
        archivo_excel = openpyxl.load_workbook(ruta_base + "\\" + n_carpeta + "\\" + n_carpeta + subcarpeta + "\\"
                                            + n_carpeta + " " + nexo + tipo + extension, data_only = True)

        lista_hojas = archivo_excel.get_sheet_names()

        index = 2
        for n_hoja in lista_hojas:
        #
            hoja = archivo_excel[n_hoja]

            #print(f"Hoja = {n_hoja} valor = {hoja[celda_promedio].value}")

            archivo_excel_objetivo = openpyxl.load_workbook(ruta_base + "\\" + archivo_objetivo)

            lista_hojas_objetivo = archivo_excel_objetivo.get_sheet_names()

            n_hoja_objetivo = lista_hojas_objetivo[index_hojas_objetivo]

            hoja_objetivo = archivo_excel_objetivo[n_hoja_objetivo]

            celda = celdas_objetivo[index_celda_obj]+str(index)
            valor_a_guardar = hoja[celda_promedio].value

            hoja_objetivo[celda].value = valor_a_guardar

            #celda = hoja_objetivo[celdas_objetivo[index_celda_obj]+str(index)]
            #celda = hoja_objetivo.cell(row=index, column=celdas_objetivo[index_celda_obj])
            #celda.value = hoja[celda_promedio].value

            archivo_excel_objetivo.save(ruta_base + "\\" + archivo_objetivo)

            index += 1

        index_celda_obj += 1

        archivo_excel_objetivo.close()
        archivo_excel.close()
        #
    index_hojas_objetivo += 1
    #
#
